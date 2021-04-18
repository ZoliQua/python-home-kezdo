import openpyxl as xl
from openpyxl.chart import LineChart, Reference

# import win32com.client
import PIL
from PIL import ImageGrab, Image
import os
import sys

import random
import datetime
import matplotlib.pyplot as plt

######## Generate automated excel workbook ########

workbook = xl.load_workbook('data/Book1.xlsx')
sheet_1 = workbook['Sheet1']

for row in range(2, sheet_1.max_row + 1):
	current = sheet_1.cell(row, 2)
	voltage = sheet_1.cell(row, 3)
	power = float(current.value) * float(voltage.value)
	power_cell = sheet_1.cell(row, 1)
	power_cell.value = power

values = Reference(sheet_1, min_row=2, max_row=sheet_1.max_row, min_col=1, max_col=1)
chart = LineChart()
chart.y_axis.title = 'Power'
chart.x_axis.title = 'Index'
chart.add_data(values)
sheet_1.add_chart(chart, 'e2')

workbook.save('data/Book1-export.xlsx')

######## Extract chart image from Excel workbook ########

# input_file = "data/Book1.xlsx"
# output_image = "images/chart.png"
#
# operation = win32com.client.Dispatch("Excel.Application")
# operation.Visible = 0
# operation.DisplayAlerts = 0
#
# workbook_2 = operation.Workbooks.Open(input_file)
# sheet_2 = operation.Sheets(1)
#
# for x, chart in enumerate(sheet_2.Shapes):
# 	chart.Copy()
# 	image = ImageGrab.grabclipboard()
# 	image.save(output_image, 'png')
# 	pass
#
# workbook_2.Close(True)
# operation.Quit()
